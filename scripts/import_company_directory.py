from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "inn_okved.xlsx"
OUTPUT = ROOT / "data" / "company-directory.json"


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def main() -> None:
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)

    directory = []
    seen = set()
    for row in rows:
        record = {
            "company_name": clean(row[1]),
            "inn": clean(row[2]),
            "okved": clean(row[3]),
            "okved_title": clean(row[4]),
            "contact_name": clean(row[5]),
            "contact_position": clean(row[6]),
            "email": clean(row[7]).lower(),
            "greeting": clean(row[8]),
        }
        if not record["company_name"] or "@" not in record["email"]:
            continue
        if record["email"] in seen:
            continue
        seen.add(record["email"])
        directory.append(record)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(directory, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "input": str(INPUT),
        "output": str(OUTPUT),
        "rows": len(directory),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
