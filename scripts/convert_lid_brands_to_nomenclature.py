from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "lid_without_calc_brand.xlsx"
OUTPUT = ROOT / "data" / "nomenclature-lid-supplement.json"


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("_x000D_", " ").strip()


def main() -> None:
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)

    items = []
    for offer_id, num, notes_ru, brand_name in rows:
        article = clean(num)
        brand = clean(brand_name)
        description = clean(notes_ru)
        if not brand or not article:
            continue
        items.append({
            "ID сделки": clean(offer_id),
            "Артикул": article,
            "Бренд": brand,
            "Наименование": description[:180],
            "Описание": description,
            "Кол-во": 1,
        })

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "input": str(INPUT),
        "output": str(OUTPUT),
        "rows": len(items),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
